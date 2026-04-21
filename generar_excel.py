# -*- coding: utf-8 -*-
"""
generar_excel.py
Usage: python generar_excel.py input.json output.xlsx

Generates a 6-sheet Excel workbook from Planify dashboard JSON data.
Requires: openpyxl
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Helper colours / fills
# ---------------------------------------------------------------------------
DARK_FILL   = PatternFill("solid", fgColor="1E293B")
ORANGE_FILL = PatternFill("solid", fgColor="FFF7ED")
ORANGE_BORDER_FILL = PatternFill("solid", fgColor="F97316")
TEAL_FILL   = PatternFill("solid", fgColor="F0FDFA")
GREEN_FILL  = PatternFill("solid", fgColor="DCFCE7")
YELLOW_FILL = PatternFill("solid", fgColor="FEF9C3")
GRAY_ALT    = PatternFill("solid", fgColor="F1F5F9")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
YELLOW_LIGHT = PatternFill("solid", fgColor="FFFDE7")
AMBER_FILL  = PatternFill("solid", fgColor="FFFBEB")
DARK_SLATE  = PatternFill("solid", fgColor="334155")

ORANGE_FONT = Font(color="F97316", bold=True)
WHITE_FONT  = Font(color="FFFFFF", bold=True)
TEAL_FONT   = Font(color="0D9488", bold=True)
GREEN_FONT  = Font(color="16A34A", bold=True)
YELLOW_FONT = Font(color="CA8A04", bold=True)
DARK_FONT   = Font(color="1E293B", bold=True)
GRAY_FONT   = Font(color="64748B")
MUTED_FONT  = Font(color="94A3B8", size=7)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_LEFT   = Alignment(horizontal="left", vertical="top", wrap_text=True)

def thin_border(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_left_border(color="F97316"):
    thick = Side(style="thick", color=color)
    thin  = Side(style="thin", color="E2E8F0")
    return Border(left=thick, right=thin, top=thin, bottom=thin)

def set_col_widths(ws, widths_chars):
    """Set column widths given list of widths in characters."""
    for i, w in enumerate(widths_chars, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def safe(val):
    """Return value or empty string if None."""
    return val if val is not None else ""

def truncate(text, max_len=25):
    s = str(text or "")
    return s[:max_len] + "..." if len(s) > max_len else s

# ---------------------------------------------------------------------------
# SHEET 1 — Resumen Ejecutivo
# ---------------------------------------------------------------------------
def build_sheet1(wb, data):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_properties.tabColor = "F97316"
    ws.sheet_view.showGridLines = False

    rango = data.get("rango", {})
    periodo_ant = data.get("periodoAnterior", {})
    kpis = data.get("kpis", {})
    mensajes = data.get("mensajes", {})

    # Column layout: [1.5, 18, 3, 18, 3, 18, 3, 18, 3, 18, 1.5]
    col_widths = [1.5, 18, 3, 18, 3, 18, 3, 18, 3, 18, 1.5]
    set_col_widths(ws, col_widths)
    # 11 columns total
    TOTAL_COLS = 11

    def merge_full(row, fill=None, height=None):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
        cell = ws.cell(row=row, column=1)
        if fill:
            cell.fill = fill
        if height:
            ws.row_dimensions[row].height = height
        return cell

    # --- ROW 1: Dark header ---
    ws.row_dimensions[1].height = 36
    merge_full(1, DARK_FILL)
    # PLANIFY in orange, subtitle after
    ws.cell(row=1, column=1).value = "PLANIFY  |  Control de Mantenimiento"
    ws.cell(row=1, column=1).font = Font(color="F97316", bold=True, size=22)
    ws.cell(row=1, column=1).alignment = LEFT

    # --- ROW 2: Orange period banner ---
    ws.row_dimensions[2].height = 22
    merge_full(2, ORANGE_FILL)
    period_text = f"Periodo: {rango.get('desde','?')} al {rango.get('hasta','?')}  |  Anterior: {periodo_ant.get('desde','?')} al {periodo_ant.get('hasta','?')}  |  Dias analizados: {rango.get('diasAnalizados','?')}"
    ws.cell(row=2, column=1).value = period_text
    ws.cell(row=2, column=1).font = Font(color="C2410C", size=9)
    ws.cell(row=2, column=1).alignment = LEFT
    # thin orange border on full row
    cell2 = ws.cell(row=2, column=1)
    cell2.border = Border(
        top=Side(style="thin", color="F97316"),
        bottom=Side(style="thin", color="F97316")
    )

    # --- ROW 3: Spacer ---
    ws.row_dimensions[3].height = 8

    # --- ROWS 4-7: KPI cards (2 rows of 4) ---
    # KPI definitions: (key, title, subtitle_key, delta_key)
    kpi_defs_row1 = [
        ("trabajosCerrados",    "Trabajos Cerrados",     "porDia",      "extra"),
        ("mediciones",          "Mediciones",            "porDia",      None),
        ("hhRegistradas",       "HH Registradas",        "porTrabajo",  None),
        ("otCerradas",          "OT Cerradas",           "avisos",      None),
    ]
    kpi_defs_row2 = [
        ("equiposIntervenidos", "Equipos Intervenidos",  None,          None),
        ("personalParticipante","Personal Participante", None,          None),
        ("lecturasCriticas",    "Lecturas Criticas",     None,          None),
        ("seguimientoActivo",   "Seguimiento Activo",    None,          None),
    ]

    # Card column starts (1-indexed): 2, 4, 6, 8 (each card spans 2 cols)
    card_cols = [2, 4, 6, 8]

    def write_kpi_card(top_row, col, kpi_key, title, subtitle_key, delta_key):
        kpi = kpis.get(kpi_key, {})
        actual = kpi.get("actual", 0)
        anterior = kpi.get("anterior", 0)
        subtitle = kpi.get(subtitle_key, "") if subtitle_key else ""
        delta_val = kpi.get(delta_key, "") if delta_key else ""

        # Compute delta if not provided
        if not delta_val and actual is not None and anterior is not None:
            diff = (actual or 0) - (anterior or 0)
            delta_val = f"+{diff}" if diff >= 0 else str(diff)

        is_positive = str(delta_val).startswith("+") if delta_val else True

        # Merge 2 columns for each card row
        for r in range(top_row, top_row + 4):
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
            c = ws.cell(row=r, column=col)
            c.fill = WHITE_FILL
            c.border = thick_left_border()

        # tag row
        tag_cell = ws.cell(row=top_row, column=col)
        tag_cell.value = title.upper()
        tag_cell.font = MUTED_FONT
        tag_cell.alignment = LEFT
        ws.row_dimensions[top_row].height = 13

        # value row
        val_cell = ws.cell(row=top_row+1, column=col)
        val_cell.value = actual
        val_cell.font = Font(color="1E293B", bold=True, size=20)
        val_cell.alignment = LEFT
        ws.row_dimensions[top_row+1].height = 28

        # subtitle row
        sub_cell = ws.cell(row=top_row+2, column=col)
        sub_cell.value = str(subtitle) if subtitle else f"anterior: {anterior}"
        sub_cell.font = Font(color="64748B", size=7)
        sub_cell.alignment = LEFT
        ws.row_dimensions[top_row+2].height = 13

        # delta badge row
        delta_cell = ws.cell(row=top_row+3, column=col)
        delta_cell.value = str(delta_val) if delta_val else ""
        delta_cell.font = GREEN_FONT if is_positive else YELLOW_FONT
        delta_cell.fill = GREEN_FILL if is_positive else YELLOW_FILL
        delta_cell.alignment = CENTER
        ws.row_dimensions[top_row+3].height = 14

    # Row 4 cards
    for i, (kpi_key, title, sub_key, delta_key) in enumerate(kpi_defs_row1):
        write_kpi_card(4, card_cols[i], kpi_key, title, sub_key, delta_key)

    # Row 8 cards (4 rows: 8,9,10,11)
    for i, (kpi_key, title, sub_key, delta_key) in enumerate(kpi_defs_row2):
        write_kpi_card(8, card_cols[i], kpi_key, title, sub_key, delta_key)

    # Spacer
    ws.row_dimensions[12].height = 8

    # --- ROWS 13+: Message boxes (4 boxes, 3 rows each) ---
    msg_keys = [
        ("lecturaEjecutiva",   "Lectura Ejecutiva"),
        ("riesgoTecnico",      "Riesgo Tecnico"),
        ("movimientoSugerido", "Movimiento Sugerido"),
        ("siguienteAccion",    "Siguiente Accion"),
    ]
    msg_base_row = 13
    msg_cols = [2, 4, 6, 8]

    for i, (msg_key, default_title) in enumerate(msg_keys):
        msg = mensajes.get(msg_key, {})
        titulo = msg.get("titulo", default_title)
        body   = msg.get("body", "Sin informacion")
        col    = msg_cols[i]

        for r in range(msg_base_row, msg_base_row + 3):
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
            c = ws.cell(row=r, column=col)
            c.fill = ORANGE_FILL
            c.border = thick_left_border()

        # Tag row
        MSG_LABELS = {
            "lecturaEjecutiva":   "LECTURA EJECUTIVA",
            "riesgoTecnico":      "RIESGO TECNICO",
            "movimientoSugerido": "MOVIMIENTO SUGERIDO",
            "siguienteAccion":    "SIGUIENTE ACCION",
        }
        tag_cell = ws.cell(row=msg_base_row, column=col)
        tag_cell.value = MSG_LABELS.get(msg_key, msg_key.upper())
        tag_cell.font = MUTED_FONT
        tag_cell.alignment = LEFT
        ws.row_dimensions[msg_base_row].height = 13

        # Title row
        title_cell = ws.cell(row=msg_base_row+1, column=col)
        title_cell.value = titulo
        title_cell.font = Font(color="C2410C", bold=True, size=10)
        title_cell.alignment = WRAP_LEFT
        ws.row_dimensions[msg_base_row+1].height = 32

        # Body row
        body_cell = ws.cell(row=msg_base_row+2, column=col)
        body_cell.value = body
        body_cell.font = Font(color="1E293B", size=8)
        body_cell.alignment = WRAP_LEFT
        ws.row_dimensions[msg_base_row+2].height = 38

        # Ensure exact tag row height
        ws.row_dimensions[msg_base_row].height = 13

    # Spacer row after messages
    spacer_row = msg_base_row + 3
    ws.row_dimensions[spacer_row].height = 8

    # --- Comparison table ---
    table_start = spacer_row + 1
    indicators = [
        ("Trabajos Cerrados",    kpis.get("trabajosCerrados",{}).get("actual",""), kpis.get("trabajosCerrados",{}).get("anterior","")),
        ("Mediciones",           kpis.get("mediciones",{}).get("actual",""),        kpis.get("mediciones",{}).get("anterior","")),
        ("HH Registradas",       kpis.get("hhRegistradas",{}).get("actual",""),     kpis.get("hhRegistradas",{}).get("anterior","")),
        ("OT Cerradas",          kpis.get("otCerradas",{}).get("actual",""),         kpis.get("otCerradas",{}).get("anterior","")),
        ("Equipos Intervenidos", kpis.get("equiposIntervenidos",{}).get("actual",""),kpis.get("equiposIntervenidos",{}).get("anterior","")),
        ("Personal Participante",kpis.get("personalParticipante",{}).get("actual",""),kpis.get("personalParticipante",{}).get("anterior","")),
        ("Lecturas Criticas",    kpis.get("lecturasCriticas",{}).get("actual",""),   kpis.get("lecturasCriticas",{}).get("anterior","")),
        ("Seguimiento Activo",   kpis.get("seguimientoActivo",{}).get("actual",""),  kpis.get("seguimientoActivo",{}).get("anterior","")),
    ]

    headers = ["Indicador", "Periodo Actual", "Periodo Anterior", "Variacion"]
    header_cols = [2, 4, 6, 8]
    header_widths = [2, 2, 2, 2]  # 2-col spans

    # Header row
    ws.row_dimensions[table_start].height = 20
    ws.merge_cells(start_row=table_start, start_column=2, end_row=table_start, end_column=TOTAL_COLS-1)
    hdr = ws.cell(row=table_start, column=2)
    hdr.value = "COMPARATIVA DE PERIODO"
    hdr.fill = DARK_FILL
    hdr.font = WHITE_FONT
    hdr.alignment = CENTER

    col_headers = ["Indicador", "Actual", "Anterior", "Variacion"]
    for j, (h, col) in enumerate(zip(col_headers, header_cols)):
        ws.merge_cells(start_row=table_start+1, start_column=col, end_row=table_start+1, end_column=col+1)
        c = ws.cell(row=table_start+1, column=col)
        c.value = h
        c.fill = DARK_FILL
        c.font = WHITE_FONT
        c.alignment = CENTER
    ws.row_dimensions[table_start+1].height = 18

    for idx, (ind, actual, anterior) in enumerate(indicators):
        row = table_start + 2 + idx
        fill = GRAY_ALT if idx % 2 == 0 else WHITE_FILL
        ws.row_dimensions[row].height = 16

        diff = ""
        try:
            diff_val = float(actual or 0) - float(anterior or 0)
            diff = f"+{int(diff_val)}" if diff_val >= 0 else str(int(diff_val))
        except Exception:
            pass

        for j, (val, col) in enumerate(zip([ind, actual, anterior, diff], header_cols)):
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            c = ws.cell(row=row, column=col)
            c.value = safe(val)
            c.fill = fill
            c.font = Font(color="1E293B", size=9)
            c.alignment = CENTER if j > 0 else LEFT
            c.border = thin_border()

    return ws


# ---------------------------------------------------------------------------
# SHEET 2 — Distribuciones
# ---------------------------------------------------------------------------
def build_sheet2(wb, data):
    ws = wb.create_sheet("Distribuciones")
    ws.sheet_properties.tabColor = "0D9488"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 28, 10, 8, 28, 2])

    dist = data.get("distribucion", {})
    lideres = data.get("lideres", [])

    row = 1

    def write_section_header(label):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1)
        c.value = label
        c.fill = DARK_FILL
        c.font = WHITE_FONT
        c.alignment = LEFT
        ws.row_dimensions[row].height = 20
        row += 1

    def write_table_header(cols):
        nonlocal row
        for j, (label, col_idx) in enumerate(zip(cols, range(2, 2 + len(cols)))):
            c = ws.cell(row=row, column=col_idx)
            c.value = label
            c.fill = PatternFill("solid", fgColor="0F766E")
            c.font = WHITE_FONT
            c.alignment = CENTER
            c.border = thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

    def write_bar_table(items, name_key="nombre", count_key="total", pct_key="pct"):
        nonlocal row
        if not items:
            ws.cell(row=row, column=2).value = "Sin datos"
            row += 1
            return
        max_total = max((item.get(count_key, 0) or 0) for item in items) or 1
        write_table_header(["Nombre", "Total", "%", "Distribucion"])
        for idx, item in enumerate(items):
            fill = TEAL_FILL if idx % 2 == 0 else WHITE_FILL
            nombre = str(item.get(name_key, ""))
            total  = item.get(count_key, 0) or 0
            pct    = item.get(pct_key, 0) or 0
            bar_len = int((total / max_total) * 20)
            bar = chr(0x2588) * bar_len  # block character

            for col_idx, val in zip(range(2, 6), [nombre, total, f"{pct}%", bar]):
                c = ws.cell(row=row, column=col_idx)
                c.value = val
                c.fill = fill
                c.font = Font(color="0D9488" if col_idx == 5 else "1E293B", size=9)
                c.alignment = LEFT if col_idx == 2 else CENTER
                c.border = thin_border()
            ws.row_dimensions[row].height = 16
            row += 1
        row += 1  # spacer

    # Especialidad
    write_section_header("DISTRIBUCION POR ESPECIALIDAD")
    write_bar_table(dist.get("especialidad", []))

    # Unidad
    write_section_header("DISTRIBUCION POR UNIDAD")
    write_bar_table(dist.get("unidad", []))

    # Condicion
    write_section_header("ESTADO DE CONDICION")
    write_bar_table(dist.get("condicion", []), name_key="estado")

    # Mix mediciones
    write_section_header("MIX DE MEDICIONES")
    write_bar_table(dist.get("mixMediciones", []), name_key="tipo")

    # Lideres
    write_section_header("RANKING DE LIDERES")
    write_table_header(["Ranking", "Lider", "Trabajos", "Unidades"])
    for idx, lider in enumerate(lideres):
        fill = TEAL_FILL if idx % 2 == 0 else WHITE_FILL
        nombre   = str(lider.get("nombre", ""))
        trabajos = lider.get("trabajos", 0)
        unidades = lider.get("unidades", 0)
        ranking_text = str(idx + 1) + "."

        for col_idx, val in zip(range(2, 6), [ranking_text, nombre, trabajos, unidades]):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill
            c.font = Font(color="1E293B", size=9)
            c.alignment = CENTER if col_idx != 3 else LEFT
            c.border = thin_border()
        ws.row_dimensions[row].height = 16
        row += 1

    return ws


# ---------------------------------------------------------------------------
# SHEET 3 — Alertas y Seguimiento
# ---------------------------------------------------------------------------
def build_sheet3(wb, data):
    ws = wb.create_sheet("Alertas y Seguimiento")
    ws.sheet_properties.tabColor = "F59E0B"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 18, 14, 14, 18, 14, 14, 2])

    alertas = data.get("alertas", [])

    # Banner row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    banner = ws.cell(row=1, column=1)
    banner.value = f"ALERTAS Y SEGUIMIENTO ACTIVO  |  {len(alertas)} registros"
    banner.fill = AMBER_FILL
    banner.font = Font(color="B45309", bold=True, size=11)
    banner.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # Spacer
    ws.row_dimensions[2].height = 8

    # Table header
    headers = ["Categoria", "Equipo", "Unidad", "Punto", "Valor", "Fecha"]
    for j, (h, col_idx) in enumerate(zip(headers, range(2, 8))):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.fill = DARK_FILL
        c.font = WHITE_FONT
        c.alignment = CENTER
        c.border = thin_border()
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = "A3"

    for idx, alerta in enumerate(alertas):
        row = 4 + idx
        fill = YELLOW_LIGHT if idx % 2 == 0 else WHITE_FILL
        vals = [
            str(alerta.get("categoria", "")),
            str(alerta.get("equipo", "")),
            str(alerta.get("unidad", "")),
            str(alerta.get("punto", "")),
            str(alerta.get("valor", "")),
            str(alerta.get("fecha", "")),
        ]
        for col_idx, val in zip(range(2, 8), vals):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill
            c.font = Font(color="1E293B", size=9)
            c.alignment = CENTER
            c.border = thin_border()
        ws.row_dimensions[row].height = 16

    if not alertas:
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=7)
        c = ws.cell(row=4, column=2)
        c.value = "Sin alertas activas en el periodo"
        c.fill = YELLOW_LIGHT
        c.font = GRAY_FONT
        c.alignment = CENTER

    return ws


# ---------------------------------------------------------------------------
# SHEET 4 — Trabajos Cerrados
# ---------------------------------------------------------------------------
def build_sheet4(wb, data):
    ws = wb.create_sheet("Trabajos Cerrados")
    ws.sheet_properties.tabColor = "334155"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 12, 14, 14, 16, 20, 12, 8, 2])

    trabajos = data.get("trabajos", [])

    # Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    banner = ws.cell(row=1, column=1)
    banner.value = f"TRABAJOS CERRADOS  |  {len(trabajos)} registros"
    banner.fill = DARK_FILL
    banner.font = WHITE_FONT
    banner.alignment = CENTER
    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 8

    # Header
    headers = ["Fecha", "Unidad", "Equipo", "Especialidad", "Lider", "OT", "HH"]
    col_map = list(range(2, 9))
    for h, col_idx in zip(headers, col_map):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.fill = DARK_FILL
        c.font = WHITE_FONT
        c.alignment = CENTER
        c.border = thin_border()
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = "B5"

    total_hh = 0.0
    for idx, trabajo in enumerate(trabajos):
        row = 4 + idx
        fill = GRAY_ALT if idx % 2 == 0 else WHITE_FILL
        especialidad = str(trabajo.get("especialidad", ""))
        hh_val = trabajo.get("hh", 0) or 0
        total_hh += float(hh_val)

        # Specialty badge fill
        if "vibr" in especialidad.lower():
            esp_fill = ORANGE_FILL
            esp_font = Font(color="F97316", size=9, bold=True)
        elif "termo" in especialidad.lower():
            esp_fill = TEAL_FILL
            esp_font = Font(color="0D9488", size=9, bold=True)
        else:
            esp_fill = fill
            esp_font = Font(color="1E293B", size=9)

        vals = [
            str(trabajo.get("fecha", "")),
            str(trabajo.get("unidad", "")),
            str(trabajo.get("equipo", "")),
            especialidad,
            str(trabajo.get("lider", "")),
            str(trabajo.get("ot", "")),
            hh_val,
        ]
        for j, (val, col_idx) in enumerate(zip(vals, col_map)):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            if j == 3:  # Especialidad column
                c.fill = esp_fill
                c.font = esp_font
            else:
                c.fill = fill
                c.font = Font(color="1E293B", size=9)
            c.alignment = CENTER if j != 4 else LEFT
            c.border = thin_border()
        ws.row_dimensions[row].height = 16

    # Totals row
    total_row = 4 + len(trabajos)
    ws.row_dimensions[total_row].height = 18
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=7)
    totals_label = ws.cell(row=total_row, column=2)
    totals_label.value = "TOTAL HH"
    totals_label.fill = DARK_FILL
    totals_label.font = WHITE_FONT
    totals_label.alignment = Alignment(horizontal="right", vertical="center")

    hh_total_cell = ws.cell(row=total_row, column=8)
    # SUM formula
    if len(trabajos) > 0:
        hh_start = 4
        hh_end = 3 + len(trabajos)
        hh_total_cell.value = f"=SUM(H{hh_start}:H{hh_end})"
    else:
        hh_total_cell.value = 0
    hh_total_cell.fill = DARK_FILL
    hh_total_cell.font = WHITE_FONT
    hh_total_cell.alignment = CENTER
    hh_total_cell.number_format = "0.0"

    return ws


# ---------------------------------------------------------------------------
# SHEET 5 — Mediciones
# ---------------------------------------------------------------------------
def build_sheet5(wb, data):
    ws = wb.create_sheet("Mediciones")
    ws.sheet_properties.tabColor = "334155"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 12, 12, 14, 14, 18, 10, 10, 10, 2])

    # Use alertas as proxy for watch measures, plus we get raw measures from trabajos
    # The JSON "alertas" field has the alert measures; we display first 30
    alertas = data.get("alertas", [])
    # Actually use the alertas list as measurement display (the schema maps this)
    mediciones_display = alertas[:30]

    total_count = len(alertas)

    # Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    banner = ws.cell(row=1, column=1)
    banner.value = f"MEDICIONES (primeras 30 de {total_count} alertas/seguimiento)"
    banner.fill = DARK_SLATE
    banner.font = WHITE_FONT
    banner.alignment = CENTER
    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 8

    # Header
    headers = ["Fecha", "Tipo", "Unidad", "Equipo", "Punto", "Valor", "Estado"]
    col_map = list(range(2, 9))
    for h, col_idx in zip(headers, col_map):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.fill = DARK_SLATE
        c.font = WHITE_FONT
        c.alignment = CENTER
        c.border = thin_border()
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = "A3"

    for idx, alerta in enumerate(mediciones_display):
        row = 4 + idx
        fill = GRAY_ALT if idx % 2 == 0 else WHITE_FILL
        categoria = str(alerta.get("categoria", ""))

        # Type badge
        if "vibr" in categoria.lower() or "vibr" in str(alerta.get("valor","")).lower():
            tipo_fill = ORANGE_FILL
            tipo_font = Font(color="F97316", size=9, bold=True)
            tipo_text = "Vibracion"
        elif "termo" in categoria.lower():
            tipo_fill = TEAL_FILL
            tipo_font = Font(color="0D9488", size=9, bold=True)
            tipo_text = "Termografia"
        else:
            tipo_fill = fill
            tipo_font = Font(color="1E293B", size=9)
            tipo_text = categoria

        estado = str(alerta.get("categoria", "Seguimiento"))
        if "criti" in estado.lower():
            estado_fill = PatternFill("solid", fgColor="FEE2E2")
            estado_font = Font(color="DC2626", size=9, bold=True)
            estado_text = "Critica"
        elif "segui" in estado.lower():
            estado_fill = YELLOW_FILL
            estado_font = Font(color="CA8A04", size=9, bold=True)
            estado_text = "Seguimiento"
        else:
            estado_fill = GREEN_FILL
            estado_font = Font(color="16A34A", size=9, bold=True)
            estado_text = "Normal"

        row_vals = [
            str(alerta.get("fecha", "")),
            tipo_text,
            str(alerta.get("unidad", "")),
            str(alerta.get("equipo", "")),
            str(alerta.get("punto", "")),
            str(alerta.get("valor", "")),
            estado_text,
        ]
        for j, (val, col_idx) in enumerate(zip(row_vals, col_map)):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            if j == 1:  # Tipo
                c.fill = tipo_fill
                c.font = tipo_font
            elif j == 6:  # Estado
                c.fill = estado_fill
                c.font = estado_font
            else:
                c.fill = fill
                c.font = Font(color="1E293B", size=9)
            c.alignment = CENTER
            c.border = thin_border()
        ws.row_dimensions[row].height = 16

    # Note at bottom
    note_row = 4 + len(mediciones_display) + 1
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=8)
    note = ws.cell(row=note_row, column=2)
    note.value = f"Total mediciones en periodo: {total_count}"
    note.font = Font(color="64748B", size=8, italic=True)
    note.alignment = LEFT

    return ws


# ---------------------------------------------------------------------------
# SHEET 6 — Ranking Operacional
# ---------------------------------------------------------------------------
def build_sheet6(wb, data):
    ws = wb.create_sheet("Ranking Operacional")
    ws.sheet_properties.tabColor = "EA580C"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 10, 20, 10, 14, 2, 2, 10, 20, 10, 14, 2])

    top_equipos = data.get("topEquipos", [])
    lideres     = data.get("lideres", [])

    # Header banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    banner = ws.cell(row=1, column=1)
    banner.value = "RANKING OPERACIONAL"
    banner.fill = PatternFill("solid", fgColor="EA580C")
    banner.font = WHITE_FONT
    banner.alignment = CENTER
    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 8

    # Left side: Top Equipos header
    equipo_headers = ["Ranking", "Equipo", "Trabajos", "Unidad"]
    for j, (h, col_idx) in enumerate(zip(equipo_headers, range(2, 6))):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.fill = DARK_FILL
        c.font = WHITE_FONT
        c.alignment = CENTER
        c.border = thin_border()
    ws.row_dimensions[3].height = 18

    # Right side: Top Lideres header (offset 5 columns: col 7+)
    lider_headers = ["Ranking", "Lider", "Trabajos", "Unidades"]
    for j, (h, col_idx) in enumerate(zip(lider_headers, range(8, 12))):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.fill = DARK_FILL
        c.font = WHITE_FONT
        c.alignment = CENTER
        c.border = thin_border()

    # Fill equipos
    for idx, equipo in enumerate(top_equipos):
        row = 4 + idx
        fill = ORANGE_FILL if idx % 2 == 0 else WHITE_FILL
        nombre   = str(equipo.get("nombre", ""))
        trabajos = equipo.get("trabajos", 0)
        unidad   = str(equipo.get("unidad", ""))
        ranking_text = str(idx + 1)

        for col_idx, val in zip(range(2, 6), [ranking_text, nombre, trabajos, unidad]):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill
            c.font = Font(color="C2410C" if col_idx == 2 else "1E293B", bold=(col_idx==2), size=9)
            c.alignment = CENTER if col_idx != 3 else LEFT
            c.border = thin_border()
        ws.row_dimensions[row].height = 16

    # Fill lideres
    for idx, lider in enumerate(lideres):
        row = 4 + idx
        fill = ORANGE_FILL if idx % 2 == 0 else WHITE_FILL
        nombre   = str(lider.get("nombre", ""))
        trabajos = lider.get("trabajos", 0)
        unidades = lider.get("unidades", 0)
        ranking_text = str(idx + 1)

        for col_idx, val in zip(range(8, 12), [ranking_text, nombre, trabajos, unidades]):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill
            c.font = Font(color="C2410C" if col_idx == 8 else "1E293B", bold=(col_idx==8), size=9)
            c.alignment = CENTER if col_idx != 9 else LEFT
            c.border = thin_border()
        ws.row_dimensions[row].height = 16

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 3:
        print("Usage: python generar_excel.py input.json output.xlsx", file=sys.stderr)
        sys.exit(1)

    input_path  = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    build_sheet1(wb, data)
    build_sheet2(wb, data)
    build_sheet3(wb, data)
    build_sheet4(wb, data)
    build_sheet5(wb, data)
    build_sheet6(wb, data)

    wb.save(output_path)
    print(f"Excel guardado en: {output_path}")


if __name__ == "__main__":
    main()
